import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2
from psycopg2.extras import execute_batch
import time
from datetime import date
import os
from dotenv import load_dotenv, find_dotenv

# Load environment variables from .env file (searching upwards if necessary)
load_dotenv(find_dotenv())

# ============================================================
# KONFIGURASI
# ============================================================
USERNAME  = os.getenv("STOCKBIT_USERNAME")
PASSWORD  = os.getenv("STOCKBIT_PASSWORD")
PLAYER_ID = os.getenv("STOCKBIT_PLAYER_ID")

# =========================
# CONFIG DATABASE
# =========================
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "database": os.getenv("DB_NAME", "postgres"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD"),
    "port": int(os.getenv("DB_PORT", 5432))
}

LOGIN_HEADERS = {
    "Content-Type": "application/json",
    "Origin": "https://stockbit.com",
    "Referer": "https://stockbit.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
}

FETCH_HEADERS_BASE = {
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate, br",
    "Origin": "https://stockbit.com",
    "Referer": "https://stockbit.com/",
    "User-Agent": "Mozilla/5.0",
}

# =========================
# CONNECT DB
# =========================
def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# =========================
# GET BROKER LIST FROM DB
# =========================
def get_brokers():
    conn = get_connection()
    cur = conn.cursor()

    query = """
    SELECT b.kode, namaperusahaan 
    FROM idxsaham.broker b 
    order by b.nilai desc;
    """

    cur.execute(query)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    # convert ke list dict
    brokers = []
    for row in rows:
        brokers.append({
            "broker_code": row[0],
            "broker_name": row[1]
        })

    return brokers

# =========================
# GET BROKER LIST FROM DB
# =========================
def get_workdays():
    conn = get_connection()
    cur = conn.cursor()

    query = """
    select trading_date  from idxsaham.trading_calendar
    where  trading_date = current_date-1
        and is_trading_day = true
        -- and trading_date > to_date('20260407','yyyymmdd')
    order by trading_date ;
    """

    cur.execute(query)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    # convert ke list dict
    workdays = []
    for row in rows:
        workdays.append({
            "work": row[0]
        })

    return workdays


# =========================
# INSERT / UPSERT DATA BROKERS ACTIVITY
# =========================
def insert_data_brokers(data):
    query = """
    INSERT INTO idxsaham.broker_activity (
        kodesaham,
        kodebroker,
        tipebroker,
        tanggal,
        nilairp,
        lot,
        avgprice,
        frekuensi,
        aksi
    )
    VALUES (%(kodesaham)s, %(kodebroker)s, %(tipebroker)s, %(tanggal)s, %(nilairp)s, 
        %(lot)s, %(avgprice)s, %(frekuensi)s, %(aksi)s)
    ON CONFLICT (tanggal, kodesaham,kodebroker)
    DO NOTHING;
    """

    # print(query)

    conn = get_connection()
    cur = conn.cursor()

    # set schema aktif
    #cur.execute("SET search_path TO idxsaham;")

    execute_batch(cur, query, data)

    conn.commit()
    cur.close()
    conn.close()

    # =========================
# INSERT / UPSERT DATA INSIDER ACTIVITY
# =========================
def insert_data_insider(data):
    query = """
    INSERT INTO idxsaham.insider_activity (
        idtrx,
        nama ,
        saham ,
        tanggal ,
        aksi ,
        sebelumnya ,
        sebelumnyapersen ,
        sekarang ,
        sekarangpersen ,
        perubahan ,
        perubahanpersen ,
        harga ,
        sumber ,
        kewarganegaraan ,
        broker ,
        badge
    )
    VALUES (%(idtrx)s, %(nama)s, %(saham)s, %(tanggal)s, %(aksi)s, %(sebelumnya)s, %(sebelumnyapersen)s, %(sekarang)s, %(sekarangpersen)s,
        %(perubahan)s, %(perubahanpersen)s, %(harga)s, %(sumber)s, %(kewarganegaraan)s, %(broker)s, %(badge)s)
    ON CONFLICT DO NOTHING;
    """

    # print(query)

    conn = get_connection()
    cur = conn.cursor()

    # set schema aktif
    #cur.execute("SET search_path TO idxsaham;")

    execute_batch(cur, query, data)

    conn.commit()
    cur.close()
    conn.close()

# ============================================================
# LOGIN
# ============================================================
def login():
    resp = requests.post(
        "https://exodus.stockbit.com/login/v6/username",
        headers=LOGIN_HEADERS,
        json={
            "user": USERNAME,
            "password": PASSWORD,
            "recaptcha_version": "RECAPTCHA_VERSION_3",
            "player_id": PLAYER_ID,
        },
    )
    if resp.status_code != 200:
        raise Exception(f"Login gagal ({resp.status_code}): {resp.text}")
    return resp.json()["data"]["login"]["token_data"]["access"]["token"]


# ============================================================
# FETCH BROKER ACTIVITY
# ============================================================
def fetch_broker_activity(token, broker_code, date_from, date_to, pages,
                          transaction_type, market_board, investor_type):
    headers = {**FETCH_HEADERS_BASE, "Authorization": f"Bearer {token}"}
    params_base = {
        "broker_code": broker_code,
        "limit": 100,
        "from": date_from,
        "to": date_to,
        "transaction_type": transaction_type,
        "market_board": market_board,
        "investor_type": investor_type,
    }

    buy_records  = []
    sell_records = []

    for page in range(1, pages + 1):
        resp = requests.get(
            "https://exodus.stockbit.com/order-trade/broker/activity",
            headers=headers,
            params={**params_base, "page": page},
        )
        if resp.status_code != 200:
            raise Exception(f"Fetch page {page} gagal ({resp.status_code}): {resp.text}")

        data = resp.json().get("data", {}).get("broker_activity_transaction", {})

        def parse_items(items, aksi):
            result = []
            for item in items:
                result.append({
                    "kodesaham": item.get("stock_code"),
                    "kodebroker": item.get("broker_code"),
                    "tipebroker": item.get("type", "").replace("BROKER_TYPE_", ""),
                    "tanggal": item.get("date"),
                    "nilairp": abs(item.get("value", 0)),
                    "lot": abs(item.get("lot", 0)),
                    "avgprice": round(item.get("avg_price", 0), 2),
                    "frekuensi": abs(item.get("freq", 0)),
                    "aksi": aksi,
                })
            return result

        buys  = data.get("brokers_buy", [])
        sells = data.get("brokers_sell", [])

        buy_records.extend(parse_items(buys, "BUY"))
        sell_records.extend(parse_items(sells, "SELL"))

        # broker activity tidak ada is_more, stop jika data < limit
        if len(buys) == 0 and len(sells) == 0:
            break
        if len(buys) < 50 and len(sells) < 50:
            break

    return buy_records, sell_records

# ============================================================
# FETCH MAJORHOLDER
# ============================================================
def fetch_majorholder(token, date_start, date_end, pages):
    BASE_URL = "https://exodus.stockbit.com/insider/company/majorholder"
    PARAMS_BASE = {
        "date_start": date_start,
        "date_end": date_end,
        "limit": pages,
        "action_type": "ACTION_TYPE_UNSPECIFIED",
        "source_type": "SOURCE_TYPE_UNSPECIFIED",
    }
    HEADERS = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Accept-Encoding": "gzip, deflate, br",
        "Origin": "https://stockbit.com",
        "Referer": "https://stockbit.com/",
        "User-Agent": "Mozilla/5.0",
    }

    records = []

    for page in range(1, 20):
        params = {**PARAMS_BASE, "page": page}
        response = requests.get(BASE_URL, headers=HEADERS, params=params)

        if response.status_code != 200:
            print(f"[Page {page}] Error {response.status_code}: {response.text}")
            break

        data = response.json()
        movements = data.get("data", {}).get("movement", [])

        if not movements:
            print(f"[Page {page}] Tidak ada data, berhenti.")
            break

        print(f"[Page {page}] Berhasil mengambil {len(movements)} data")

        for item in movements:
            records.append({
                "idtrx": item.get("id"),
                "nama": item.get("name"),
                "saham": item.get("symbol"),
                "tanggal": item.get("date"),
                "aksi": item.get("action_type", "").replace("ACTION_TYPE_", ""),
                "sebelumnya": item.get("previous", {}).get("value"),
                "sebelumnyapersen": item.get("previous", {}).get("percentage"),
                "sekarang": item.get("current", {}).get("value"),
                "sekarangpersen": item.get("current", {}).get("percentage"),
                "perubahan": item.get("changes", {}).get("value"),
                "perubahanpersen": item.get("changes", {}).get("percentage"),
                "harga": item.get("price_formatted"),
                "sumber": item.get("data_source", {}).get("label"),
                "kewarganegaraan": item.get("nationality", "").replace("NATIONALITY_TYPE_", ""),
                "broker": item.get("broker_detail", {}).get("code"),
                "badge": ", ".join(
                    [b.replace("SHAREHOLDER_BADGE_", "") for b in item.get("badges", [])]
                ),
            })

            for x in records:
                print(x)

        if not data.get("data", {}).get("is_more", False):
            break

    return records

# =========================
# MAIN EXECUTION
# =========================
if __name__ == "__main__":
    token = login()

    brokers = get_brokers()
    workdays = get_workdays()

    today = date.today()
    num = 1

    for work in workdays:
        date_from        = work["work"]
        date_to          = today
        pages            = 10

    #    records = fetch_majorholder(token, date_from, date_to, pages)
    #    insert_data_insider(records)
    #    print(f"\n✅ Data insider berhasil disimpan!", {date_from})
              
        for broker in brokers:
            broker_code      = broker["broker_code"]
            transaction_type = "TRANSACTION_TYPE_NET"
            market_board     = "MARKET_TYPE_REGULER"
            investor_type    = "INVESTOR_TYPE_ALL"

            buy_records, sell_records = fetch_broker_activity(
                    token, broker_code, date_from, date_to, pages,
                    transaction_type, market_board, investor_type,
                )
            insert_data_brokers(buy_records)
            insert_data_brokers(sell_records)
            print(f"\n✅ {num}. Data broker summary berhasil disimpan!  {date_from}", [broker_code])
            num = num+1

            # delay biar ga kena rate limit
            time.sleep(1)
